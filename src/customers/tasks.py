# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font
from django.conf import settings
from django.db.models.fields.files import FieldFile
from celery.task import task


@task
def queryset_to_xlsx_response(query_set, response):
    def get_field_value(obj, field_name):
        value = getattr(obj, field_name)
        if issubclass(value.__class__, FieldFile):
            if value:
                return settings.SITE_URL + settings.MEDIA_URL + value.name
            else:
                return u''

        return value

    model_meta = query_set.model._meta

    work_book = openpyxl.Workbook()
    work_sheet = work_book.get_active_sheet()
    work_sheet.title = model_meta.verbose_name

    row = 0
    model_fields = [
        field for field in model_meta.concrete_fields
    ]

    for col, field in enumerate(model_fields):
        cell = work_sheet.cell(row=row + 1, column=col + 1)
        cell.value = field.verbose_name
        cell.font = Font(bold=True)

    for query_obj in query_set:
        row += 1

        for col, field in enumerate(model_fields):
            cell = work_sheet.cell(row=row + 1, column=col + 1)
            cell.value = get_field_value(query_obj, field.name)

    work_book.save(response)

    return response
